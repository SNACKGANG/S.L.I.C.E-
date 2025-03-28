import openpyxl
import discord

from io import BytesIO


class RoleService:
    async def process_roles_from_excel(self, ctx, role_id: int, file_bytes: bytes, action: str):
        workbook = openpyxl.load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
            member_id = row[0]
            if member_id is None:
                continue

            try:
                member_id = int(member_id)
            except ValueError:
                print(f"Invalid ID: {member_id}")
                continue

            member = ctx.guild.get_member(member_id)
            if not member:
                await ctx.send(f"Member with ID {member_id} not found.")
                continue

            role = ctx.guild.get_role(role_id)
            if not role:
                await ctx.send(f"Role with ID {role_id} not found.")
                return

            if action == "add":
                await self._add_role(ctx, member, role)
            elif action == "remove":
                await self._remove_role(ctx, member, role)

    @staticmethod
    async def _add_role(ctx, member: discord.Member, role: discord.Role):
        if role not in member.roles:
            await member.add_roles(role)
            await ctx.send(f"Role added to {member.display_name}.")
        else:
            await ctx.send(f"{member.display_name} already has the role")

    @staticmethod
    async def _remove_role(ctx, member: discord.Member, role: discord.Role):
        if role in member.roles:
            await member.remove_roles(role)
            await ctx.send(f"Role removed from {member.display_name}.")
        else:
            await ctx.send(f"{member.display_name} does not have the role.")
